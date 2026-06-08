import re
import unicodedata
from pathlib import Path
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = Path("neu_cfs_big.xlsx")
OUTPUT_FILE = Path("neu_data_labeled.xlsx")

# ========== 1. TEENCODE & VIẾT TẮT (ĐẦY ĐỦ NHẤT) ==========
TEEN_CODE_MAP = {
    # ---------- CƠ BẢN ----------
    "k": "không", "ko": "không", "kh": "không", "kp": "không phải",
    "kphai": "không phải", "k0": "không",
    "cx": "cũng", "cug": "cũng", "cg": "cũng",
    "thui": "thôi", "thoy": "thôi", "thoi": "thôi",
    "mk": "mình", "mik": "mình", "mikc": "mình cũng",
    "mn": "mọi người", "mng": "mọi người", "mọing": "mọi người",
    "ng": "người",
    "r": "rồi", "rùi": "rồi", "roi": "rồi",
    "z": "vậy", "zay": "vậy", "vay": "vậy",
    "w": "với", "vs": "với",
    "x": "không",
    "cc": "cực kỳ", "vcl": "quá",
    "cak": "các", "cn": "công nhận", "nma": "nhưng mà", "nhx": "nhưng", "nx": "nữa",
    "bth": "bình thường", "bthg": "bình thường", "bt": "bình thường",
    "dc": "được", "đc": "được",
    "dk": "đăng ký", "đk": "đăng ký", "đki": "đăng ký", "dkhp": "đăng ký học phần",
    "ad": "admin", "ib": "nhắn tin", "rep": "trả lời", "nt": "nhắn tin", "trl": "trả lời",
    "cfs": "confession",
    "ntn": "như thế nào", "tn": "thế nào",
    "kbt": "không biết", "kb": "không biết",
    "tkb": "thời khóa biểu",
    "gpa": "điểm trung bình", "cre": "tín chỉ", "sub": "môn học",
    "đrl": "điểm rèn luyện", "drl": "điểm rèn luyện",
    "onl": "online", "off": "nghỉ", "out": "ra ngoài",
    "onsite": "trực tiếp", "online": "trực tuyến",
    "pass": "qua môn", "fail": "trượt", "hok": "học", "hk": "học kỳ",
    "tl": "tài liệu", "kt": "kiểm tra", "av": "anh văn", "nl": "năng lượng",
    
    # ---------- TỪ KHÓA NEU ĐẶC THÙ ----------
    "sv": "sinh viên", "svien": "sinh viên", "sv năm nhất": "sinh viên năm nhất",
    "gv": "giảng viên", "gvien": "giảng viên",
    "ql": "quản lý", "qly": "quản lý", "nv": "nhân viên",
    "ktx": "ký túc xá", "tx": "túc xá",
    "aep": "chương trình tiên tiến", "clc": "chất lượng cao",
    "tmktqt": "thương mại và kinh tế quốc tế", "tmu": "thương mại",
    
    # ---------- SỬA LỖI CHÍNH TẢ ----------
    "thủ thuộc": "thuộc", "thu": "thuộc", "thuoc": "thuộc", "thuôc": "thuộc",
    "thủ cơ": "thủ công", "thủ công": "thủ công",
    "hong": "không", "khong": "không",
    "duoc": "được", "dang": "đang", "diem": "điểm",
    "thilai": "thi lại", "thi lai": "thi lại",
    "skien": "sự kiện", "sk": "sự kiện",
    "tui": "tôi", "cho tui xin": "cho tôi xin",
    "m": "mình", "b": "bạn", "e": "em", "a": "anh", "c": "chị",
    "thi": "thì", "thik": "thích", "thich": "thích",
    "ko": "không", "hog": "không", "hok": "không",
}

# ========== 2. TỪ KHÓA PHÂN LOẠI CHỦ ĐỀ ==========
# 2a. CƠ SỞ VẬT CHẤT - NHẬN XÉT về phòng ốc, thiết bị, tiện nghi
CSVC_KEYWORDS = {
    # Phòng học, giảng đường
    "phòng học", "giảng đường", "hội trường", "lớp học", "phòng máy",
    # Thiết bị
    "máy chiếu", "máy lạnh", "điều hòa", "quạt", "đèn", "loa", "máy tính", "máy in", "bảng", "phấn",
    # Tiện nghi
    "wifi", "mạng", "internet", "nhà vệ sinh", "thư viện", "căng tin", "bãi xe", "thang máy",
    "ký túc xá", "ktx", "nước uống", "ổ điện", "bục giảng",
    # Tình trạng
    "nóng", "lạnh", "bẩn", "sạch", "mất điện", "hỏng", "xuống cấp", "ngột ngạt", "ồn", "tối",
    # Khu vực
    "a1", "a2", "b1", "b2", "c1", "c2", "cơ sở vật chất", "trang thiết bị", "quét mặt"
}

# 2b. CHẤT LƯỢNG GIẢNG DẠY - NHẬN XÉT về giảng viên, bài giảng, thi cử
CLGD_KEYWORDS = {
    # Giảng viên
    "giảng viên", "thầy", "cô", "giáo viên", "trợ giảng",
    "dạy dở", "dạy hay", "dạy tốt", "dạy chán", "dạy nhanh", "dạy chậm",
    "giảng dở", "giảng hay", "giảng khó hiểu", "giảng dễ hiểu",
    "nhiệt tình", "vô trách nhiệm", "hỗ trợ", "phản hồi", "hướng dẫn",
    # Bài giảng, nội dung
    "bài giảng", "slide", "giáo trình", "tài liệu", "nội dung",
    "bài tập", "bài tập nhóm", "thuyết trình", "đồ án", "khóa luận", "tiểu luận",
    # Thi cử, điểm số
    "đề thi", "đề khó", "đề dễ", "đề dài", "điểm", "chấm điểm", "chấm thi",
    "thi cử", "kiểm tra", "điểm danh", "qua môn", "trượt môn", "phúc khảo",
    "điểm thấp", "điểm cao", "điểm kém", "thi lại", "cải thiện", "học cải thiện",
    # Chất lượng đào tạo
    "chất lượng giảng dạy", "đào tạo kém", "đào tạo tốt", "chương trình học", "học phí",
    "học được", "học không được", "học online", "học trực tuyến"
}

# 2c. HỎI THÔNG TIN - KHÔNG PHẢI NHẬN XÉT
INFO_KEYWORDS = {
    "lịch thi", "lịch học", "thời khóa biểu", "tkb", "lịch kiểm tra",
    "đăng ký ở đâu", "nộp ở đâu", "làm ở đâu", "lấy ở đâu", "xin ở đâu",
    "bao giờ", "khi nào", "mấy giờ", "tháng mấy", "ngày nào",
    "thủ tục", "hồ sơ", "giấy tờ", "đơn từ", "xin xác nhận",
    "học phí bao nhiêu", "giá bao nhiêu", "bao nhiêu tiền",
    "liên hệ ai", "hỏi ai", "gặp ai", "ai biết", "có ai biết"
}

# ========== 3. TỪ KHÓA CẢM XÚC ==========
EMOTION_KEYWORDS = {
    "Hai_Long": {
        "tốt", "ổn", "ok", "oke", "ổn áp", "hài lòng", "thích", "tuyệt",
        "xịn", "nhiệt tình", "ổn định", "sạch", "đẹp", "hay", "được",
        "thuận tiện", "thoải mái", "dễ chịu", "hỗ trợ", "tận tình",
        "nhanh", "hiệu quả", "rõ ràng", "dễ hiểu", "chất lượng", "ưng",
        "chu đáo", "xuất sắc", "tuyệt vời", "quá tốt", "rất tốt"
    },
    "Khong_Hai_Long": {
        "tệ", "kém", "chán", "bức xúc", "nóng", "bẩn", "ồn", "lag", "lỗi",
        "không hài lòng", "khó chịu", "mất", "hỏng", "phiền", "không ổn",
        "không tốt", "chậm", "khó", "áp lực", "căng thẳng", "xuống cấp",
        "thiếu", "bất tiện", "trễ", "mệt", "rắc rối", "tức", "bực", "dở",
        "thất vọng", "quá tệ", "ức chế", "bực mình"
    }
}

# ========== 4. DỮ LIỆU CẦN LOẠI BỎ ==========
GARBAGE_NAMES = {
    # Tên người dùng
    "memorablehamster3324", "tet_binh_ngo_7754", "cavin0301", "positivebee8241",
    "calmdeer1047", "wildcat1085", "lazylion7445", "happyfrog2768", "smarthawk8354",
    "boldpanda6631", "smartlion8930", "proudwolf8552", "happyfox7530",
    "cleverpanda5871", "happytiger2269", "lazylion3428", "happyrabbit2326",
    "bravedeer2805", "coolfox8329", "busytiger3942", "cleverrabbit2825",
    "calmdeer9547", "proudbear5917", "cleverdog7391", "happytiger2326",
    # Tên người thật
    "phuong linh", "hoang anh tuyet", "hoàng ánh tuyết", "kim chi",
    "mai trang", "trang anh", "quyn dinh", "vu thi lan", "le hai dang",
    "le van phuc", "nguyen minh quan", "nguyen thu ha", "pham bao chau",
    "tran khanh linh", "tran thi bich", "duong van hung", "ho hai yen",
    "mai thi huong", "bui minh khoa", "bui thi hoa", "hoang ngoc anh",
    "vu minh quan", "phan thi mai", "dang thanh hang", "le thuy duong",
    "tran minh anh", "do quoc nam", "nguyen ha linh", "pham thuy linh"
}

GARBAGE_PHRASES = {
    "công khai", "tham gia", "theo dõi", "quản trị viên", "người kiểm duyệt",
    "tác giả", "xem tất cả", "đã chỉnh sửa", "đang hoạt động", "phản hồi",
    "bình luận", "thích", "trả lời", "chia sẻ", "ảnh đại diện", "xem thêm"
}

# ========== 5. HÀM XỬ LÝ ==========
def normalize_text(text: str) -> str:
    """Chuẩn hóa Unicode (bỏ lỗi font)"""
    text = unicodedata.normalize("NFC", text)
    fixes = {
        "à": "à", "á": "á", "ả": "ả", "ã": "ã", "ạ": "ạ",
        "ằ": "ằ", "ắ": "ắ", "ẳ": "ẳ", "ẵ": "ẵ", "ặ": "ặ",
        "ầ": "ầ", "ấ": "ấ", "ẩ": "ẩ", "ẫ": "ẫ", "ậ": "ậ",
        "è": "è", "é": "é", "ẻ": "ẻ", "ẽ": "ẽ", "ẹ": "ẹ",
        "ề": "ề", "ế": "ế", "ể": "ể", "ễ": "ễ", "ệ": "ệ",
        "ì": "ì", "í": "í", "ỉ": "ỉ", "ĩ": "ĩ", "ị": "ị",
        "ò": "ò", "ó": "ó", "ỏ": "ỏ", "õ": "õ", "ọ": "ọ",
        "ồ": "ồ", "ố": "ố", "ổ": "ổ", "ỗ": "ỗ", "ộ": "ộ",
        "ờ": "ờ", "ớ": "ớ", "ở": "ở", "ỡ": "ỡ", "ợ": "ợ",
        "ù": "ù", "ú": "ú", "ủ": "ủ", "ũ": "ũ", "ụ": "ụ",
        "ừ": "ừ", "ứ": "ứ", "ử": "ử", "ữ": "ữ", "ự": "ự",
        "ỳ": "ỳ", "ý": "ý", "ỷ": "ỷ", "ỹ": "ỹ", "ỵ": "ỵ",
        "Đ": "Đ", "đ": "đ"
    }
    for old, new in fixes.items():
        text = text.replace(old, new)
    return text

def expand_teen_code(text: str) -> str:
    """Mở rộng teencode"""
    result = text
    # Sắp xếp từ dài trước để thay thế chính xác
    for key, val in sorted(TEEN_CODE_MAP.items(), key=lambda x: len(x[0]), reverse=True):
        # Thay thế từ độc lập
        result = re.sub(rf'\b{re.escape(key)}\b', val, result, flags=re.IGNORECASE)
        # Thay thế liền kề
        result = re.sub(rf'{re.escape(key)}', val, result, flags=re.IGNORECASE)
    return result

def is_valid_content(text: str) -> bool:
    """Kiểm tra nội dung có hợp lệ không"""
    if not text or len(text) < 15:
        return False
    
    text_lower = text.lower().strip()
    
    # Loại bỏ tên người, username
    if text_lower in GARBAGE_NAMES:
        return False
    
    # Loại bỏ username dạng chữ+số
    if re.match(r"^[a-z0-9_]+$", text_lower) and len(text) < 30:
        return False
    
    # Loại bỏ tên người dạng "Hoàng Ánh Tuyết"
    if re.match(r"^[A-ZÀ-Ỹ][a-zà-ỹ]+(?:\s+[A-ZÀ-Ỹ][a-zà-ỹ]+){0,2}$", text):
        if len(text.split()) <= 3:
            return False
    
    # Loại bỏ phrase rác
    if text_lower in GARBAGE_PHRASES:
        return False
    
    return True

def clean_text(value: object) -> str:
    """Làm sạch văn bản hoàn chỉnh"""
    if pd.isna(value):
        return ""
    
    text = normalize_text(str(value))
    
    # 1. Xóa URL và link
    text = re.sub(r"https?://\S+|www\.\S+", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"/groups/\S+", " ", text, flags=re.IGNORECASE)
    
    # 2. Xóa pattern Facebook
    text = re.sub(r"Người\s+tham\s+gia\s+ẩn\s+danh\s*\d*", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d+\s*(?:giây|phút|giờ|ngày|tuần|tháng|năm)\b", " ", text, flags=re.IGNORECASE)
    
    # 3. Xóa emoji
    text = re.sub(r"[\U0001F004-\U0001F9FF\u2600-\u26FF\u2700-\u27BF]+", " ", text, flags=re.UNICODE)
    
    # 4. Xóa ký tự đặc biệt (giữ chữ, số, dấu câu cơ bản)
    text = re.sub(r"[^0-9A-Za-zÀ-ỹ\s?!.,]", " ", text, flags=re.UNICODE)
    
    # 5. Xử lý lặp ký tự
    text = re.sub(r"(.)\1{3,}", r"\1\1", text)
    text = re.sub(r'([!?.])\1+', r'\1', text)
    
    # 6. Mở rộng teencode
    text = expand_teen_code(text)
    
    # 7. Xóa phrase rác
    for phrase in GARBAGE_PHRASES:
        text = re.sub(rf'\b{re.escape(phrase)}\b', "", text, flags=re.IGNORECASE)
    
    # 8. Chuẩn hóa khoảng trắng
    text = re.sub(r"\s+", " ", text).strip()
    text = text.strip(".,!?;:")
    
    # 9. Kiểm tra hợp lệ
    if not is_valid_content(text):
        return ""
    
    return text

def classify_topic(text: str) -> str:
    """
    PHÂN LOẠI CHỦ ĐỀ - QUY TẮC RÕ RÀNG:
    1. Cơ sở vật chất: có từ khóa CSVC và KHÔNG có từ khóa CLGD
    2. Chất lượng giảng dạy: có từ khóa CLGD
    3. Hỏi thông tin: câu hỏi về lịch, địa điểm, thủ tục
    4. Không đủ dữ liệu: không xác định được
    """
    if not text:
        return "Không đủ dữ liệu"
    
    text_lower = text.lower()
    
    # ----- KIỂM TRA HỎI THÔNG TIN -----
    is_asking = False
    for kw in INFO_KEYWORDS:
        if kw in text_lower:
            is_asking = True
            break
    
    if is_asking and "?" in text_lower:
        return "Hỏi thông tin"
    
    # ----- KIỂM TRA CHẤT LƯỢNG GIẢNG DẠY -----
    clgd_matches = [kw for kw in CLGD_KEYWORDS if kw in text_lower]
    if len(clgd_matches) >= 1:
        return "Chất lượng giảng dạy"
    
    # ----- KIỂM TRA CƠ SỞ VẬT CHẤT -----
    csvc_matches = [kw for kw in CSVC_KEYWORDS if kw in text_lower]
    if len(csvc_matches) >= 1:
        return "Cơ sở vật chất"
    
    # ----- MẶC ĐỊNH -----
    return "Không đủ dữ liệu"

def classify_satisfaction(text: str, topic: str = None) -> str:
    """
    PHÂN LOẠI MỨC ĐỘ HÀI LÒNG
    - Hài lòng: từ khóa tích cực > từ khóa tiêu cực
    - Không hài lòng: từ khóa tiêu cực > từ khóa tích cực
    - Trung lập: số lượng bằng nhau hoặc không rõ ràng
    - Không đủ dữ liệu: không thể xác định
    """
    if not text:
        return "Không đủ dữ liệu"
    
    # Nếu là câu hỏi thông tin, không có cảm xúc
    if topic == "Hỏi thông tin":
        return "Không đủ dữ liệu"
    
    text_lower = text.lower()
    
    positive_score = 0
    negative_score = 0
    
    # Đếm từ khóa
    for word in EMOTION_KEYWORDS["Hai_Long"]:
        if word in text_lower:
            positive_score += 1
    
    for word in EMOTION_KEYWORDS["Khong_Hai_Long"]:
        if word in text_lower:
            negative_score += 1
    
    # Từ khóa mạnh (tăng trọng số)
    strong_positive = ["tuyệt", "xuất sắc", "tuyệt vời", "quá tốt", "rất tốt", "rất hài lòng"]
    strong_negative = ["tệ hại", "thất vọng", "quá tệ", "rất tệ", "cực kỳ tệ"]
    
    for word in strong_positive:
        if word in text_lower:
            positive_score += 2
    
    for word in strong_negative:
        if word in text_lower:
            negative_score += 2
    
    # Xử lý phủ định (không tốt = không hài lòng)
    negations = ["không", "chẳng", "chả", "chưa", "đéo"]
    words = text_lower.split()
    for i, word in enumerate(words):
        if word in negations and i + 1 < len(words):
            if words[i + 1] in EMOTION_KEYWORDS["Hai_Long"]:
                positive_score -= 1
                negative_score += 1
    
    # Kết luận
    if positive_score == 0 and negative_score == 0:
        return "Không đủ dữ liệu"
    
    if negative_score > positive_score:
        return "Không hài lòng"
    elif positive_score > negative_score:
        return "Hài lòng"
    else:
        return "Trung lập"

def classify_sentence_type(text: str) -> str:
    """Phân loại câu hỏi / câu khẳng định"""
    if not text:
        return "Không xác định"
    
    question_words = ("ai", "ở đâu", "sao", "tại sao", "vì sao", "bao giờ", 
                      "khi nào", "bao nhiêu", "như thế nào", "thế nào", "có ai", 
                      "cho mình hỏi", "cho em hỏi", "mọi người ơi", "ai biết")
    
    text_lower = text.lower().strip()
    
    if "?" in text_lower:
        return "Câu hỏi"
    
    if text_lower.startswith(question_words):
        return "Câu hỏi"
    
    return "Câu khẳng định"

def classify_record_type(record_type: object) -> str:
    """Phân loại bài viết / bình luận"""
    if pd.isna(record_type):
        return "Bài viết"
    text = str(record_type).lower()
    if "comment" in text:
        return "Bình luận"
    return "Bài viết"

# ========== 6. STYLE EXCEL CHUYÊN NGHIỆP ==========
def style_excel_output(file_path: Path) -> None:
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Màu header
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    
    # Màu chủ đề
    topic_colors = {
        "Cơ sở vật chất": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "Chất lượng giảng dạy": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "Hỏi thông tin": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "Không đủ dữ liệu": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    }
    
    # Màu cảm xúc
    satisfaction_colors = {
        "Hài lòng": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        "Không hài lòng": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "Trung lập": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        "Không đủ dữ liệu": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    }
    
    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD")
    )
    
    # Style header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Lấy vị trí cột
    headers = [cell.value for cell in ws[1]]
    topic_idx = headers.index("Chủ đề") if "Chủ đề" in headers else -1
    sat_idx = headers.index("Mức độ hài lòng") if "Mức độ hài lòng" in headers else -1
    
    # Style dữ liệu
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if topic_idx >= 0:
            topic = row[topic_idx].value
            if topic in topic_colors:
                row[topic_idx].fill = topic_colors[topic]
        
        if sat_idx >= 0:
            sat = row[sat_idx].value
            if sat in satisfaction_colors:
                row[sat_idx].fill = satisfaction_colors[sat]
        
        ws.row_dimensions[row_idx].height = 45
    
    # Độ rộng cột
    col_widths = {"A": 15, "B": 55, "C": 70, "D": 24, "E": 20, "F": 15}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = "A2"
    wb.save(file_path)

# ========== 7. HÀM CHÍNH ==========
def main():
    print("="*70)
    print("🚀 PHÂN LOẠI DỮ LIỆU CONFESSION NEU")
    print("   Chủ đề: CSVC | CLGD | Hỏi thông tin | Không đủ dữ liệu")
    print("   Cảm xúc: Hài lòng | Không hài lòng | Trung lập | Không đủ dữ liệu")
    print("="*70)
    
    if not INPUT_FILE.exists():
        print(f"❌ Không tìm thấy file: {INPUT_FILE}")
        return
    
    print(f"\n📂 Đang đọc file: {INPUT_FILE.name}")
    df = pd.read_excel(INPUT_FILE)
    print(f"📊 Tổng số dòng: {len(df)}")
    
    content_col = df.columns[2]
    record_col = df.columns[0]
    
    print("\n🧹 Bước 1: Làm sạch dữ liệu...")
    df["Nội dung sạch"] = df[content_col].apply(clean_text)
    
    df_valid = df[df["Nội dung sạch"] != ""].copy()
    removed = len(df) - len(df_valid)
    print(f"   ✅ Giữ lại: {len(df_valid)} dòng")
    print(f"   🗑️ Loại bỏ: {removed} dòng ({removed/len(df)*100:.1f}%)")
    
    print("\n🏷️ Bước 2: Phân loại chủ đề...")
    df_valid["Chủ đề"] = df_valid["Nội dung sạch"].apply(classify_topic)
    
    print("😊 Bước 3: Phân loại cảm xúc...")
    df_valid["Mức độ hài lòng"] = df_valid.apply(
        lambda row: classify_satisfaction(row["Nội dung sạch"], row["Chủ đề"]), axis=1
    )
    
    print("❓ Bước 4: Phân loại loại câu...")
    df_valid["Loại câu"] = df_valid["Nội dung sạch"].apply(classify_sentence_type)
    
    print("📌 Bước 5: Phân loại bản ghi...")
    df_valid["Loại bản ghi"] = df_valid[record_col].apply(classify_record_type)
    df_valid["Nội dung gốc"] = df_valid[content_col]
    
    # Xuất file
    output_cols = ["Loại bản ghi", "Nội dung gốc", "Nội dung sạch", "Chủ đề", "Mức độ hài lòng", "Loại câu"]
    output_df = df_valid[output_cols].copy()
    
    print(f"\n💾 Bước 6: Xuất file {OUTPUT_FILE.name}...")
    output_df.to_excel(OUTPUT_FILE, index=False)
    
    print("🎨 Bước 7: Định dạng Excel...")
    style_excel_output(OUTPUT_FILE)
    
    # ========== THỐNG KÊ CHI TIẾT ==========
    print("\n" + "="*60)
    print("📊 THỐNG KÊ KẾT QUẢ")
    print("="*60)
    
    print("\n🏷️ PHÂN LOẠI CHỦ ĐỀ:")
    topic_order = ["Cơ sở vật chất", "Chất lượng giảng dạy", "Hỏi thông tin", "Không đủ dữ liệu"]
    for topic in topic_order:
        cnt = len(df_valid[df_valid["Chủ đề"] == topic])
        if cnt > 0:
            pct = cnt/len(df_valid)*100
            bar = "█" * int(pct / 2)
            print(f"   {topic:20}: {cnt:4} dòng ({pct:5.1f}%) {bar}")
    
    print("\n😊 PHÂN LOẠI MỨC ĐỘ HÀI LÒNG:")
    sat_order = ["Hài lòng", "Không hài lòng", "Trung lập", "Không đủ dữ liệu"]
    for sat in sat_order:
        cnt = len(df_valid[df_valid["Mức độ hài lòng"] == sat])
        if cnt > 0:
            pct = cnt/len(df_valid)*100
            bar = "█" * int(pct / 2)
            print(f"   {sat:18}: {cnt:4} dòng ({pct:5.1f}%) {bar}")
    
    print("\n❓ PHÂN LOẠI LOẠI CÂU:")
    for qtype in ["Câu hỏi", "Câu khẳng định"]:
        cnt = len(df_valid[df_valid["Loại câu"] == qtype])
        if cnt > 0:
            pct = cnt/len(df_valid)*100
            bar = "█" * int(pct / 2)
            print(f"   {qtype:15}: {cnt:4} dòng ({pct:5.1f}%) {bar}")
    
    print("\n📌 PHÂN LOẠI BẢN GHI:")
    for rtype in ["Bài viết", "Bình luận"]:
        cnt = len(df_valid[df_valid["Loại bản ghi"] == rtype])
        if cnt > 0:
            pct = cnt/len(df_valid)*100
            bar = "█" * int(pct / 2)
            print(f"   {rtype:15}: {cnt:4} dòng ({pct:5.1f}%) {bar}")
    
    print(f"\n✅ HOÀN THÀNH! File lưu tại: {OUTPUT_FILE.resolve()}")
    print("="*60)

if __name__ == "__main__":
    main()